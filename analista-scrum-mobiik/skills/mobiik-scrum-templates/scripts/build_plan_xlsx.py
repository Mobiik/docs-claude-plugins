"""
Plan de Trabajo Excel — Migración Preselección TecMilenio
Hojas: Portada · Dashboard · Resumen Ejecutivo · WBS · Cronograma (Gantt) · Sprints · Recursos · Hitos · RAID (link)
Estándar visual Mobiik.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import date, timedelta

BG_DARK   = "0A0A0A"
BG_PANEL  = "252525"
LIME      = "AADC1E"
WHITE     = "FFFFFF"
GREY      = "AAAAAA"
RED       = "FF4444"
ORANGE    = "FFA040"
BLUE      = "1EA0DC"
LT_GREY   = "F5F5F5"
GANTT_BG  = "1A1A1A"

FONT = "Arial"

def fill(c): return PatternFill("solid", start_color=c, end_color=c)
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def styled_title(ws, title, n_cols):
    for col in range(1, n_cols + 2):
        ws.cell(row=1, column=col).fill = fill(BG_DARK)
        ws.cell(row=2, column=col).fill = fill(BG_DARK)
        ws.cell(row=3, column=col).fill = fill(BG_DARK)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1+n_cols)
    ws["B2"] = title
    ws["B2"].font = Font(name=FONT, size=18, bold=True, color=WHITE)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 2+n_cols):
        ws.cell(row=4, column=col).fill = fill(LIME)
    ws.row_dimensions[4].height = 3

def write_header_row(ws, headers, row):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=2+i, value=h)
        c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
        c.fill = fill(BG_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

def add_footer(ws, row, n_cols):
    ws.cell(row=row, column=2, value="© 2026 Mobiik   |   coding for the future   |   Confidencial")
    ws.cell(row=row, column=2).font = Font(name=FONT, size=9, italic=True, color=GREY)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+n_cols)


# ═══ HOJA 1: PORTADA ═══════════════════════════════════════════════
ws = wb.active
ws.title = "Portada"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
for c in 'BCDEFG':
    ws.column_dimensions[c].width = 22

styled_title(ws, "PLAN DE TRABAJO", 6)

# Title block
ws["B3"] = "MIGRACIÓN PRESELECCIÓN — TECMILENIO"
ws["B3"].font = Font(name=FONT, size=13, bold=True, color=LIME)
ws["B3"].fill = fill(BG_DARK)
ws.merge_cells("B3:G3")

# Info table
info = [
    ("Cliente",          "TecMilenio"),
    ("Proyecto",         "Migración Técnica Preselección — Lift-and-Shift a Azure PaaS"),
    ("Modalidad",        "Proyecto fijo · Metodología Scrum"),
    ("Inversión Total",  "$218,790.00 MXN (antes de IVA)"),
    ("Hito 1 — Kickoff", "$65,637.00 (30%) — Inicio Sprint 1"),
    ("Hito 2 — Cierre",  "$153,153.00 (70%) — Cierre Sprint 3"),
    ("Duración",         "3 semanas · 15 días útiles · 3 Sprints"),
    ("Fecha de inicio",  "Mayo 22, 2026"),
    ("Fecha Go-Live",    "Junio 11, 2026"),
    ("Scrum Master",     "Miroslava Jiménez (Mobiik)"),
    ("Product Owner",    "PO TecMilenio"),
    ("Proyecto Jira",    "TEC26 — Tec Migración Pre-selección"),
    ("ID Propuesta",     "Migración Preselección — 20-may-2026"),
]
for i, (k, v) in enumerate(info):
    r = 7 + i
    c1 = ws.cell(row=r, column=2, value=k)
    c1.font = Font(name=FONT, size=10, bold=True, color="000000")
    c1.fill = fill(LT_GREY); c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border = border
    c2 = ws.cell(row=r, column=3, value=v)
    c2.font = Font(name=FONT, size=10)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = border
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

# Versiones
ws["B22"] = "CONTROL DE VERSIONES"
ws["B22"].font = Font(name=FONT, size=12, bold=True, color=LIME)
ws.merge_cells("B22:G22")
write_header_row(ws, ["Versión", "Fecha", "Autor", "Cambios", "Estado", "Aprobado"], 23)
ws.row_dimensions[23].height = 28
v = ["v1.0", "25-may-2026", "Miroslava Jiménez", "Versión inicial del plan", "Borrador", "PO TecMilenio"]
for i, val in enumerate(v):
    c = ws.cell(row=24, column=2+i, value=val)
    c.font = Font(name=FONT, size=10); c.border = border
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")

add_footer(ws, 30, 6)


# ═══ HOJA 2: DASHBOARD EJECUTIVO ═══════════════════════════════════
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for c in 'BCDEFGH':
    ws.column_dimensions[c].width = 16

styled_title(ws, "DASHBOARD EJECUTIVO", 7)

# 4 KPI cards
kpis = [
    ("INVERSIÓN",  "$218,790",  "MXN — Antes de IVA"),
    ("DURACIÓN",   "15 días",   "3 semanas / 3 Sprints"),
    ("EQUIPO",     "5",         "personas (SM, Arq, Dev, DevOps, QA)"),
    ("AVANCE",     "0%",        "Proyecto no iniciado"),
]
for i, (lbl, big, sub) in enumerate(kpis):
    col = 2 + i*2
    # Label
    c = ws.cell(row=6, column=col, value=lbl)
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.fill = fill(BG_PANEL); c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
    # Big value
    c2 = ws.cell(row=7, column=col, value=big)
    c2.font = Font(name=FONT, size=22, bold=True, color=LIME)
    c2.fill = fill(BG_DARK); c2.border = border
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
    # Subtitle
    c3 = ws.cell(row=8, column=col, value=sub)
    c3.font = Font(name=FONT, size=9, color=GREY)
    c3.fill = fill(BG_DARK); c3.border = border
    c3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 38
    ws.row_dimensions[8].height = 24

# Avance por Épica
ws["B11"] = "AVANCE POR ÉPICA"
ws["B11"].font = Font(name=FONT, size=12, bold=True, color=LIME)
ws.merge_cells("B11:H11")

epic_hdr = ["#", "Épica", "Sprint", "HUs", "Completadas", "% Avance"]
write_header_row(ws, epic_hdr, 12)
ws.row_dimensions[12].height = 26

epics = [
    ("1", "Análisis Técnico y Planeación",        "S1",    4),
    ("2", "Preparación de Entornos Azure",         "S1-S2", 6),
    ("3", "Migración Lift-and-Shift",              "S2",    5),
    ("4", "Migración de Bases de Datos",           "S2",    4),
    ("5", "Pipelines CI/CD",                       "S2-S3", 3),
    ("6", "Atención de 13 Hallazgos de Seguridad", "S3",    16),  # 3 HUs + 13 subt
    ("7", "Pruebas y UAT",                         "S3",    4),
    ("8", "Go-Live y Cierre",                      "Cierre", 3),
]
for i, (n, name, sp, hus) in enumerate(epics):
    r = 13 + i
    vals = [n, name, sp, hus, 0, f"=E{r}/D{r}"]
    for j, val in enumerate(vals):
        c = ws.cell(row=r, column=2+j, value=val)
        c.font = Font(name=FONT, size=10); c.border = border
        c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        if j == 5:
            c.number_format = "0%"
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif j in (0, 2, 3, 4):
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Total row
r_total = 13 + len(epics)
ws.cell(row=r_total, column=2, value="TOTAL").font = Font(name=FONT, size=11, bold=True, color=BG_DARK)
ws.cell(row=r_total, column=2).fill = fill(LIME)
ws.cell(row=r_total, column=2).alignment = Alignment(horizontal="center")
ws.cell(row=r_total, column=2).border = border
ws.merge_cells(start_row=r_total, start_column=2, end_row=r_total, end_column=4)
ws.cell(row=r_total, column=5, value=f"=SUM(E13:E{r_total-1})").font = Font(name=FONT, size=11, bold=True)
ws.cell(row=r_total, column=5).fill = fill(LIME)
ws.cell(row=r_total, column=5).alignment = Alignment(horizontal="center")
ws.cell(row=r_total, column=5).border = border
ws.cell(row=r_total, column=6, value=f"=SUM(F13:F{r_total-1})").font = Font(name=FONT, size=11, bold=True)
ws.cell(row=r_total, column=6).fill = fill(LIME)
ws.cell(row=r_total, column=6).alignment = Alignment(horizontal="center")
ws.cell(row=r_total, column=6).border = border
ws.cell(row=r_total, column=7, value=f"=SUM(F13:F{r_total-1})/SUM(D13:D{r_total-1})").font = Font(name=FONT, size=11, bold=True, color=BG_DARK)
ws.cell(row=r_total, column=7).fill = fill(LIME)
ws.cell(row=r_total, column=7).number_format = "0%"
ws.cell(row=r_total, column=7).alignment = Alignment(horizontal="center")
ws.cell(row=r_total, column=7).border = border

# Hitos próximos
ws["B25"] = "HITOS PRÓXIMOS"
ws["B25"].font = Font(name=FONT, size=12, bold=True, color=LIME)
ws.merge_cells("B25:H25")
write_header_row(ws, ["#", "Hito", "Fecha", "Responsable", "Estado"], 26)
hitos = [
    ("1", "Kickoff oficial",                   "22-may-2026", "Mobiik × TecMilenio", "Próximo"),
    ("2", "Entrega de accesos por TecMilenio", "22-may-2026", "TecMilenio",          "Pendiente"),
    ("3", "Cierre Sprint 1 (Análisis)",         "26-may-2026", "Mobiik",              "Pendiente"),
    ("4", "Cierre Sprint 2 (Migración)",        "2-jun-2026",  "Mobiik",              "Pendiente"),
    ("5", "Cierre Sprint 3 (QA + UAT)",         "10-jun-2026", "Mobiik × TecMilenio", "Pendiente"),
    ("6", "Go-Live productivo",                "11-jun-2026", "Mobiik × TecMilenio", "Pendiente"),
]
for i, row in enumerate(hitos):
    r = 27 + i
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=2+j, value=val)
        c.font = Font(name=FONT, size=10); c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")

add_footer(ws, 36, 7)


# ═══ HOJA 3: WBS (Work Breakdown Structure) ═══════════════════════
ws = wb.create_sheet("WBS")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
widths = [10, 35, 50, 15, 12, 12, 14]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(2+i)].width = w

styled_title(ws, "WORK BREAKDOWN STRUCTURE", 7)

write_header_row(ws, ["ID WBS", "Entregable / Actividad", "Descripción", "Sprint", "Esfuerzo (días)", "Responsable", "Jira Epic"], 6)
ws.row_dimensions[6].height = 32

wbs = [
    ("1",     "ANÁLISIS Y PLANEACIÓN",                        "Fase de descubrimiento técnico",                       "S1",  "2",  "Mobiik",                "TEC26-1"),
    ("1.1",   "  Revisión arquitectura actual",                 "Inventario de componentes y dependencias",             "S1",  "0.5", "Arquitecto",           "TEC26-1"),
    ("1.2",   "  Identificación dependencias técnicas",         "Mapeo de frameworks y librerías",                       "S1",  "0.5", "DevOps",               "TEC26-1"),
    ("1.3",   "  Validación requerimientos Azure",              "Confirmar viabilidad Lift-and-Shift",                   "S1",  "0.5", "Arquitecto",           "TEC26-1"),
    ("1.4",   "  Documentación de riesgos técnicos",            "Actualización del RAID Log",                            "S1",  "0.5", "Scrum Master",         "TEC26-1"),

    ("2",     "PREPARACIÓN DE ENTORNOS AZURE",                  "Provisión y configuración de servicios PaaS",          "S1-S2", "2",  "Mobiik",              "TEC26-2"),
    ("2.1",   "  Resource Groups",                              "Creación con nomenclatura estándar",                    "S1-S2", "0.25", "DevOps",            "TEC26-2"),
    ("2.2",   "  App Services (backend)",                       "Plan + App Service + slots",                            "S2",  "0.5", "DevOps",               "TEC26-2"),
    ("2.3",   "  Key Vault",                                    "Gestión centralizada de secretos",                     "S2",  "0.5", "DevOps",               "TEC26-2"),
    ("2.4",   "  Storage Account (condicional)",                "Si aplica según análisis",                              "S2",  "0.25", "DevOps",              "TEC26-2"),
    ("2.5",   "  Azure Managed Redis (condicional)",            "Si aplica según análisis",                              "S2",  "0.25", "DevOps",              "TEC26-2"),
    ("2.6",   "  Definición de ambientes",                       "No-Prod y Prod con naming convention",                 "S2",  "0.25", "DevOps",              "TEC26-2"),

    ("3",     "MIGRACIÓN LIFT-AND-SHIFT",                       "Despliegue del aplicativo migrado",                    "S2",  "4",  "Mobiik",                "TEC26-3"),
    ("3.1",   "  Backend .NET en App Service",                  "Despliegue + healthcheck",                              "S2",  "1.5", "Dev Fullstack",        "TEC26-3"),
    ("3.2",   "  Frontend AngularJS en Static Web Apps",         "Cuando aplique",                                       "S2",  "1",  "Dev Fullstack",        "TEC26-3"),
    ("3.3",   "  Variables de entorno y secretos",              "App Settings + Key Vault",                             "S2",  "0.5", "Dev Fullstack",        "TEC26-3"),
    ("3.4",   "  Cadenas de conexión",                          "Connection strings desde Key Vault",                   "S2",  "0.5", "Dev Fullstack",        "TEC26-3"),
    ("3.5",   "  Manejo de archivos y rutas",                    "Validación I/O en cloud",                              "S2",  "0.5", "Dev Fullstack",        "TEC26-3"),

    ("4",     "MIGRACIÓN BASES DE DATOS",                       "MS SQL Server → Azure SQL",                            "S2",  "3",  "Mobiik",                "TEC26-4"),
    ("4.1",   "  Migración de esquemas y datos",                 "Backup + DMS + validación",                            "S2",  "1.5", "DevOps",               "TEC26-4"),
    ("4.2",   "  Configuración de conectividad",                 "Firewall + connection strings",                        "S2",  "0.5", "DevOps",               "TEC26-4"),
    ("4.3",   "  Validación de completitud",                     "Conteos + checksums",                                  "S2",  "0.5", "QA",                   "TEC26-4"),
    ("4.4",   "  Pruebas smoke con datos migrados",              "Validación end-to-end",                                "S2",  "0.5", "QA",                   "TEC26-4"),

    ("5",     "PIPELINES CI/CD",                                "Automatización de despliegues",                       "S2-S3", "2",  "Mobiik",              "TEC26-5"),
    ("5.1",   "  Configuración de pipelines",                   "Build + deploy automatizado",                          "S2-S3", "1.5", "DevOps",             "TEC26-5"),
    ("5.2",   "  Integración GitHub",                            "Webhooks + secrets",                                   "S2-S3", "0.25", "DevOps",            "TEC26-5"),
    ("5.3",   "  Flujos de deployment",                          "Diagrama + runbook",                                   "S3",  "0.25", "DevOps",              "TEC26-5"),

    ("6",     "ATENCIÓN DE HALLAZGOS DE SEGURIDAD",             "Mitigación de 13 hallazgos identificados",            "S3",  "4",  "Mobiik",                "TEC26-6"),
    ("6.1",   "  Hallazgos 1-7 (versiones y disclosures)",      "Actualizaciones de librerías y headers",               "S3",  "1",  "Dev/SecOps",           "TEC26-6"),
    ("6.2",   "  Hallazgos 8-9 (CSRF, jQuery disclosure)",      "Implementación de tokens y minificación",              "S3",  "1.5", "Dev/SecOps",           "TEC26-6"),
    ("6.3",   "  Hallazgos 10-13 (Bootstrap, Modernizr, etc)",  "Mitigaciones restantes",                                "S3",  "0.5", "Dev/SecOps",           "TEC26-6"),
    ("6.4",   "  Reescaneo",                                     "Validación con herramienta de seguridad",              "S3",  "0.5", "SecOps",               "TEC26-6"),
    ("6.5",   "  Validación funcional post-cambios",             "Regresión por QA",                                     "S3",  "0.5", "QA",                   "TEC26-6"),

    ("7",     "PRUEBAS Y UAT",                                   "Smoke tests internos + UAT cliente",                  "S3",  "3",  "Mobiik × TecMilenio",   "TEC26-7"),
    ("7.1",   "  Smoke tests internos",                          "Validación funcional + integraciones",                  "S3",  "0.5", "QA",                  "TEC26-7"),
    ("7.2",   "  Acompañamiento UAT",                             "3 días con equipo de pruebas del cliente",             "S3",  "1.5", "QA Mobiik + QA Tec",  "TEC26-7"),
    ("7.3",   "  Corrección de incidencias",                      "Atención de defectos críticos/altos",                  "S3",  "0.75", "Dev",                "TEC26-7"),
    ("7.4",   "  Firma de aceptación",                            "Documento firmado por PO",                             "S3",  "0.25", "Scrum Master",       "TEC26-7"),

    ("8",     "GO-LIVE Y CIERRE",                                "Despliegue productivo y cierre formal",               "Cierre", "1", "Mobiik × TecMilenio", "TEC26-8"),
    ("8.1",   "  Despliegue productivo",                         "Pipeline a Prod con aprobación",                       "Cierre", "0.5", "DevOps",            "TEC26-8"),
    ("8.2",   "  Acompañamiento post-liberación",                "Estabilización inicial",                                "Cierre", "0.25", "Scrum Master",     "TEC26-8"),
    ("8.3",   "  Carta de cierre + memoria técnica",             "Entregables finales firmados",                          "Cierre", "0.25", "Scrum Master",     "TEC26-8"),
]
for i, row in enumerate(wbs):
    r = 7 + i
    is_header = '.' not in row[0]  # main category
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=2+j, value=val)
        c.border = border
        if is_header:
            c.font = Font(name=FONT, size=11, bold=True, color=WHITE)
            c.fill = fill(BG_PANEL)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            c.font = Font(name=FONT, size=10)
            c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if j in (3, 4):  # Sprint, Esfuerzo
                c.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "C7"
add_footer(ws, 7 + len(wbs) + 2, 7)


# ═══ HOJA 4: CRONOGRAMA / GANTT ═══════════════════════════════════
ws = wb.create_sheet("Cronograma")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 8   # ID
ws.column_dimensions['C'].width = 35  # Actividad
ws.column_dimensions['D'].width = 12  # Sprint
ws.column_dimensions['E'].width = 14  # Inicio
ws.column_dimensions['F'].width = 14  # Fin
ws.column_dimensions['G'].width = 8   # Días
# Días del proyecto (15 columnas)
for i in range(15):
    ws.column_dimensions[get_column_letter(8 + i)].width = 5

styled_title(ws, "CRONOGRAMA TIPO GANTT", 21)

# Headers
write_header_row(ws, ["ID", "Actividad", "Sprint", "Inicio", "Fin", "Días"], 6)
# Day headers
project_start = date(2026, 5, 22)
labor_days = []
d = project_start
while len(labor_days) < 15:
    if d.weekday() < 5:  # 0=Mon..4=Fri
        labor_days.append(d)
    d += timedelta(days=1)

for i, day in enumerate(labor_days):
    c = ws.cell(row=6, column=8+i, value=day.strftime("%d-%b"))
    c.font = Font(name=FONT, size=8, bold=True, color=WHITE)
    c.fill = fill(BG_DARK); c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[6].height = 32

# Tasks with day ranges (start_day_idx, end_day_idx inclusive, 0-based)
tasks = [
    ("1.0", "Sprint 1: Análisis y Planeación",            "S1",      "22-may", "26-may", 3, 0, 2,   BG_PANEL),
    ("1.1", "Revisión arquitectura actual",                "S1",      "22-may", "22-may", 1, 0, 0,   LIME),
    ("1.2", "Identificación dependencias",                 "S1",      "22-may", "25-may", 2, 0, 1,   LIME),
    ("1.3", "Validación requerimientos Azure",             "S1",      "25-may", "26-may", 2, 1, 2,   LIME),
    ("1.4", "Documentación de riesgos",                    "S1",      "26-may", "26-may", 1, 2, 2,   LIME),

    ("2.0", "Preparación de Entornos Azure",               "S1-S2",   "26-may", "29-may", 4, 2, 5,   BG_PANEL),
    ("2.1", "Resource Groups + App Services",              "S2",      "27-may", "28-may", 2, 3, 4,   LIME),
    ("2.2", "Key Vault + Storage + Redis",                 "S2",      "28-may", "29-may", 2, 4, 5,   LIME),

    ("3.0", "Migración Lift-and-Shift",                    "S2",      "27-may", "1-jun",  4, 3, 6,   BG_PANEL),
    ("3.1", "Backend .NET en App Service",                 "S2",      "27-may", "28-may", 2, 3, 4,   LIME),
    ("3.2", "Frontend AngularJS",                          "S2",      "28-may", "29-may", 2, 4, 5,   LIME),
    ("3.3", "Variables, secretos, conexiones",             "S2",      "29-may", "1-jun",  2, 5, 6,   LIME),

    ("4.0", "Migración BD MS SQL → Azure SQL",             "S2",      "28-may", "1-jun",  3, 4, 6,   BG_PANEL),
    ("4.1", "Migración esquemas y datos",                  "S2",      "28-may", "29-may", 2, 4, 5,   RED),
    ("4.2", "Validación + smoke test BD",                  "S2",      "1-jun",  "1-jun",  1, 6, 6,   LIME),

    ("5.0", "Pipelines CI/CD",                             "S2-S3",   "1-jun",  "3-jun",  3, 6, 8,   BG_PANEL),
    ("5.1", "Configuración pipelines + GitHub",            "S2-S3",   "1-jun",  "2-jun",  2, 6, 7,   LIME),
    ("5.2", "Flujos de deployment",                        "S3",      "3-jun",  "3-jun",  1, 8, 8,   LIME),

    ("6.0", "Atención de 13 Hallazgos de Seguridad",        "S3",      "2-jun",  "5-jun",  4, 7, 10,  BG_PANEL),
    ("6.1", "Hallazgos críticos (CSRF, ISE, HttpOnly)",    "S3",      "2-jun",  "3-jun",  2, 7, 8,   RED),
    ("6.2", "Hallazgos de versiones y disclosures",        "S3",      "3-jun",  "4-jun",  2, 8, 9,   ORANGE),
    ("6.3", "Reescaneo + validación funcional",            "S3",      "5-jun",  "5-jun",  1, 10, 10, LIME),

    ("7.0", "Pruebas y UAT",                               "S3",      "8-jun",  "10-jun", 3, 11, 13, BG_PANEL),
    ("7.1", "Smoke tests internos",                        "S3",      "8-jun",  "8-jun",  1, 11, 11, LIME),
    ("7.2", "UAT cliente",                                 "S3",      "8-jun",  "10-jun", 3, 11, 13, BLUE),
    ("7.3", "Corrección + Firma aceptación",               "S3",      "10-jun", "10-jun", 1, 13, 13, LIME),

    ("8.0", "Go-Live y Cierre",                            "Cierre",  "11-jun", "11-jun", 1, 14, 14, BG_PANEL),
    ("8.1", "Despliegue productivo",                       "Cierre",  "11-jun", "11-jun", 1, 14, 14, LIME),
    ("8.2", "Carta de cierre + memoria técnica",            "Cierre",  "11-jun", "11-jun", 1, 14, 14, LIME),
]

for i, (id_, act, sp, ini, fin, dias, ds, de, color) in enumerate(tasks):
    r = 7 + i
    is_phase = id_.endswith(".0")

    # Info columns (B-G)
    info_vals = [id_, act, sp, ini, fin, dias]
    for j, val in enumerate(info_vals):
        c = ws.cell(row=r, column=2+j, value=val)
        c.border = border
        if is_phase:
            c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
            c.fill = fill(BG_PANEL)
        else:
            c.font = Font(name=FONT, size=9)
            c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        if j in (0, 2, 5):
            c.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Gantt bars (columns 8..22)
    for day_idx in range(15):
        col = 8 + day_idx
        c = ws.cell(row=r, column=col)
        c.border = border
        if ds <= day_idx <= de:
            c.fill = fill(color)
        else:
            c.fill = fill("FFFFFF" if not is_phase else "1A1A1A")

ws.freeze_panes = "H7"

# Legend
last_row = 7 + len(tasks) + 2
ws.cell(row=last_row, column=2, value="LEYENDA:").font = Font(name=FONT, size=10, bold=True, color=BG_DARK)
legend_items = [(LIME, "Actividad regular"), (RED, "Riesgo Alto / Crítico"),
                (ORANGE, "Riesgo Medio"), (BLUE, "UAT Cliente"),
                (BG_PANEL, "Fase / Hito mayor")]
col_pos = 3
for color, label in legend_items:
    ws.cell(row=last_row, column=col_pos).fill = fill(color)
    ws.cell(row=last_row, column=col_pos).border = border
    ws.cell(row=last_row, column=col_pos+1, value=label).font = Font(name=FONT, size=9)
    ws.merge_cells(start_row=last_row, start_column=col_pos+1, end_row=last_row, end_column=col_pos+2)
    col_pos += 3

add_footer(ws, last_row + 3, 21)


# ═══ HOJA 5: SPRINTS ═══════════════════════════════════════════════
ws = wb.create_sheet("Sprints")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for w, c in zip([12, 18, 15, 15, 35, 12, 14, 15], 'BCDEFGHI'):
    ws.column_dimensions[c].width = w

styled_title(ws, "PLANIFICACIÓN POR SPRINT", 8)

write_header_row(ws, ["Sprint", "Período", "Día Inicio", "Día Cierre", "Objetivo principal", "HUs", "Capacidad (días)", "Velocity prev."], 6)
ws.row_dimensions[6].height = 32

sprints = [
    ("Sprint 1",  "Semana 1 (L-M)",    "22-may-2026", "26-may-2026", "Análisis técnico, identificación de dependencias, arquitectura objetivo aprobada", 4,  "10 (5 personas × 2 días)", "N/A — proyecto nuevo"),
    ("Sprint 2",  "Semana 1-2 (X-L)",  "27-may-2026", "1-jun-2026",  "Lift-and-Shift backend, migración BD MS SQL, conectividad",                       17, "20 (5 personas × 4 días)", "10 (de Sprint 1)"),
    ("Sprint 3",  "Semana 2-3 (M-V)",  "2-jun-2026",  "10-jun-2026", "13 hallazgos de seguridad, smoke tests, pipelines CI/CD y UAT",                   28, "35 (5 personas × 7 días)", "17"),
    ("Go-Live",   "Semana 3 (J)",      "11-jun-2026", "11-jun-2026", "Despliegue productivo, acompañamiento post-liberación, carta de cierre",          3,  "5 (5 personas × 1 día)",   "—"),
]
for i, row in enumerate(sprints):
    r = 7 + i
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=2+j, value=val)
        c.font = Font(name=FONT, size=10); c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        if j == 0:
            c.font = Font(name=FONT, size=11, bold=True, color=BG_DARK)
            c.fill = fill(LIME)
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 38

# Velocity formula
ws.cell(row=11, column=2, value="VELOCITY ACUMULADO").font = Font(name=FONT, size=11, bold=True, color=LIME)
ws.merge_cells("B11:I11")
write_header_row(ws, ["Sprint", "HUs Planeadas", "HUs Completadas", "Velocity", "% Cumplimiento", "Notas"], 12)
for i, sp in enumerate(["Sprint 1", "Sprint 2", "Sprint 3", "Go-Live"]):
    r = 13 + i
    hus = sprints[i][5]
    vals = [sp, hus, 0, f"=D{r}", f"=D{r}/C{r}", "Por iniciar"]
    for j, val in enumerate(vals):
        c = ws.cell(row=r, column=2+j, value=val)
        c.font = Font(name=FONT, size=10); c.border = border
        c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        if j == 4:
            c.number_format = "0%"

add_footer(ws, 22, 8)


# ═══ HOJA 6: RECURSOS Y ASIGNACIONES ═══════════════════════════════
ws = wb.create_sheet("Recursos")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for w, c in zip([22, 22, 12, 12, 12, 12, 12, 14], 'BCDEFGHI'):
    ws.column_dimensions[c].width = w

styled_title(ws, "RECURSOS Y ASIGNACIONES", 8)

write_header_row(ws, ["Nombre", "Rol", "% Asignación", "S1 (días)", "S2 (días)", "S3 (días)", "Go-Live", "Total"], 6)
ws.row_dimensions[6].height = 32

recursos = [
    ("Miroslava Jiménez",  "Scrum Master",        100,   2,  4,  7,  1),
    ("Andrés Hernández",   "Arquitecto",          100,   2,  4,  4,  1),
    ("Iran Arellanes",     "Dev Fullstack",       100,   1,  4,  7,  1),
    ("Norberto Galicia",   "DevOps",              100,   2,  4,  5,  1),
    ("César Zúñiga",       "QA Tester",           100,   1,  4,  7,  1),
]
for i, row in enumerate(recursos):
    r = 7 + i
    name, role, pct, d1, d2, d3, dg = row
    total_formula = f"=SUM(E{r}:H{r})"
    vals = [name, role, pct/100, d1, d2, d3, dg, total_formula]
    for j, val in enumerate(vals):
        c = ws.cell(row=r, column=2+j, value=val)
        c.font = Font(name=FONT, size=10); c.border = border
        c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        if j == 0:
            c.font = Font(name=FONT, size=10, bold=True)
            c.alignment = Alignment(horizontal="left", vertical="center")
        if j == 2:
            c.number_format = "0%"

# Total
r_tot = 7 + len(recursos)
ws.cell(row=r_tot, column=2, value="TOTAL DÍAS").font = Font(name=FONT, size=11, bold=True, color=BG_DARK)
ws.cell(row=r_tot, column=2).fill = fill(LIME)
ws.cell(row=r_tot, column=2).alignment = Alignment(horizontal="center")
ws.cell(row=r_tot, column=2).border = border
ws.merge_cells(start_row=r_tot, start_column=2, end_row=r_tot, end_column=4)
for col in range(5, 10):
    c = ws.cell(row=r_tot, column=col, value=f'=SUM({get_column_letter(col)}7:{get_column_letter(col)}{r_tot-1})')
    c.font = Font(name=FONT, size=11, bold=True, color=BG_DARK)
    c.fill = fill(LIME); c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center")

add_footer(ws, r_tot + 3, 8)


# ═══ HOJA 7: HITOS Y ENTREGABLES ═══════════════════════════════════
ws = wb.create_sheet("Hitos")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 2
for w, c in zip([10, 38, 35, 14, 18, 18], 'BCDEFG'):
    ws.column_dimensions[c].width = w

styled_title(ws, "HITOS Y ENTREGABLES CONTRACTUALES", 6)

write_header_row(ws, ["#", "Entregable", "Criterio de aceptación", "Sprint", "Hito Facturación", "Estado"], 6)
ws.row_dimensions[6].height = 32

entregables = [
    ("1",  "Código fuente de aplicación migrada",        "Repositorio actualizado y validado",                    "S3",      "—",     "Pendiente"),
    ("2",  "Infraestructura Azure (ambiente no-prod)",   "Ambientes No-Prod operativos",                          "S1-S2",   "—",     "Pendiente"),
    ("3",  "Análisis técnico inicial",                    "Informe entregado y validado",                          "S1",      "Hito 1", "Pendiente"),
    ("4",  "Memoria técnica de cambios aplicados",        "Documento técnico revisado",                            "S3",      "Hito 2", "Pendiente"),
    ("5",  "Pipelines de CI/CD",                          "Pipelines ejecutando deploys correctamente",           "S2-S3",   "Hito 2", "Pendiente"),
    ("6",  "Especificación de CI/CD e infra Azure",       "Documentación entregada",                               "S3",      "Hito 2", "Pendiente"),
    ("7",  "Corrección de 13 hallazgos de seguridad",     "Evidencia + reescaneo limpio",                          "S3",      "Hito 2", "Pendiente"),
    ("8",  "Resultados de smoke test",                     "Sin hallazgos críticos abiertos",                       "S3",      "Hito 2", "Pendiente"),
    ("9",  "Carta de cierre del proyecto",                 "Firmada por TecMilenio",                                "Cierre",  "Hito 2", "Pendiente"),
]
for i, row in enumerate(entregables):
    r = 7 + i
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=2+j, value=val)
        c.font = Font(name=FONT, size=10); c.border = border
        c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if j in (0, 3, 4, 5):
            c.alignment = Alignment(horizontal="center", vertical="center")

# Pagos
ws.cell(row=18, column=2, value="ESQUEMA DE PAGOS").font = Font(name=FONT, size=12, bold=True, color=LIME)
ws.merge_cells("B18:G18")
write_header_row(ws, ["Pago", "Hito", "Cuándo", "%", "Monto MXN", "Estado"], 19)
pagos = [
    ("1", "Inicio del proyecto / Sprint 1",  "Inicio / Kickoff",        "30%", "$65,637.00",   "Pendiente"),
    ("2", "Migración, QA, UAT y cierre",     "Cierre Sprint 3 (Sem 3)", "70%", "$153,153.00",  "Pendiente"),
    ("",  "TOTAL",                            "—",                       "100%", "$218,790.00", "—"),
]
for i, row in enumerate(pagos):
    r = 20 + i
    is_total = row[1] == "TOTAL"
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=2+j, value=val)
        c.border = border
        if is_total:
            c.font = Font(name=FONT, size=11, bold=True, color=BG_DARK)
            c.fill = fill(LIME)
        else:
            c.font = Font(name=FONT, size=10)
            c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")

add_footer(ws, 25, 6)


# ═══ HOJA 8: ENLACE A RAID ═════════════════════════════════════════
ws = wb.create_sheet("RAID (link)")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
for c in 'BCDEF':
    ws.column_dimensions[c].width = 22

styled_title(ws, "RAID LOG — Riesgos, Supuestos, Issues, Dependencias", 5)

ws["B7"] = "📋 Para detalles completos, consulta el archivo:"
ws["B7"].font = Font(name=FONT, size=12, color=WHITE); ws.merge_cells("B7:F7")

ws["B9"] = "RAID Log - Migración Preselección.xlsx"
ws["B9"].font = Font(name=FONT, size=14, bold=True, color=LIME)
ws.merge_cells("B9:F9")

ws["B11"] = "Resumen:"
ws["B11"].font = Font(name=FONT, size=11, bold=True, color=WHITE); ws.merge_cells("B11:F11")

resumen = [
    ("Riesgos identificados",        "8"),
    ("Riesgos CRÍTICOS",              "1 (R-03: Retraso en accesos por TecMilenio)"),
    ("Riesgos ALTOS",                 "3"),
    ("Supuestos del proyecto",        "10"),
    ("Dependencias del cliente",      "9"),
    ("Dependencias CRÍTICAS",         "3 (D-01, D-02, D-03 — Accesos en Sprint 1)"),
]
for i, (k, v) in enumerate(resumen):
    r = 13 + i
    c1 = ws.cell(row=r, column=2, value=k)
    c1.font = Font(name=FONT, size=10, bold=True)
    c1.fill = fill(LT_GREY); c1.alignment = Alignment(horizontal="left", vertical="center")
    c1.border = border
    c2 = ws.cell(row=r, column=3, value=v)
    c2.font = Font(name=FONT, size=10)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = border
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)

add_footer(ws, 25, 5)


# ═══ REORDER + SAVE ════════════════════════════════════════════════
wb._sheets = [
    wb["Portada"], wb["Dashboard"], wb["WBS"], wb["Cronograma"],
    wb["Sprints"], wb["Recursos"], wb["Hitos"], wb["RAID (link)"]
]

out = "./entregables/Plan de Trabajo - Migración Preselección.xlsx"
wb.save(out)
print(f"✓ Saved: {out}")
print(f"  Sheets: {wb.sheetnames}")
