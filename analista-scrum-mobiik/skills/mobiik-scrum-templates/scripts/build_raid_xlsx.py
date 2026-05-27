"""
RAID Log — Migración Preselección TecMilenio
Estándar visual Mobiik: header oscuro #0A0A0A texto blanco Arial, acentos #AADC1E
Semaforización: rojo #FF4444 / naranja #FFA040 / lima #AADC1E
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# Brand colors (no leading #)
BG_DARK   = "0A0A0A"
BG_PANEL  = "252525"
LIME      = "AADC1E"
WHITE     = "FFFFFF"
GREY      = "AAAAAA"
RED       = "FF4444"
ORANGE    = "FFA040"
BLUE      = "1EA0DC"
LT_GREY   = "F5F5F5"

FONT = "Arial"

# Reusable styles
def hdr_font():    return Font(name=FONT, size=11, bold=True, color=WHITE)
def title_font():  return Font(name=FONT, size=18, bold=True, color=WHITE)
def body_font():   return Font(name=FONT, size=10, color="000000")
def body_white():  return Font(name=FONT, size=10, color=WHITE)
def lime_font():   return Font(name=FONT, size=10, bold=True, color=LIME)

def fill(color):   return PatternFill("solid", start_color=color, end_color=color)
def center():      return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left_wrap():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ─── HOJA 1: PORTADA / CONTROL ─────────────────────────────────────
ws = wb.active
ws.title = "Portada"
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
for c in 'BCDEFG':
    ws.column_dimensions[c].width = 22

# Dark header strip
for col in range(1, 8):
    ws.cell(row=1, column=col).fill = fill(BG_DARK)
    ws.cell(row=2, column=col).fill = fill(BG_DARK)
    ws.cell(row=3, column=col).fill = fill(BG_DARK)
    ws.cell(row=4, column=col).fill = fill(BG_DARK)
    ws.cell(row=5, column=col).fill = fill(BG_DARK)

ws.merge_cells("B2:G2"); ws["B2"] = "RAID LOG"
ws["B2"].font = title_font(); ws["B2"].fill = fill(BG_DARK); ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells("B3:G3"); ws["B3"] = "MIGRACIÓN PRESELECCIÓN — TECMILENIO"
ws["B3"].font = Font(name=FONT, size=13, bold=True, color=LIME); ws["B3"].fill = fill(BG_DARK)

ws.merge_cells("B4:G4"); ws["B4"] = "Riesgos · Supuestos · Issues · Dependencias"
ws["B4"].font = Font(name=FONT, size=11, italic=True, color=GREY); ws["B4"].fill = fill(BG_DARK)

# Lime accent bar
for col in range(2, 8):
    ws.cell(row=6, column=col).fill = fill(LIME)
ws.row_dimensions[6].height = 4

# Project info table
ws["B8"]  = "Cliente";          ws["C8"]  = "TecMilenio"
ws["B9"]  = "Proyecto";         ws["C9"]  = "Migración Técnica Preselección — Azure PaaS"
ws["B10"] = "Modalidad";        ws["C10"] = "Lift-and-Shift · Precio fijo"
ws["B11"] = "Inversión";        ws["C11"] = "$218,790.00 MXN (antes de IVA)"
ws["B12"] = "Duración";         ws["C12"] = "3 semanas · 15 días útiles · 3 Sprints"
ws["B13"] = "Scrum Master";     ws["C13"] = "Miroslava Jiménez (Mobiik)"
ws["B14"] = "Proyecto Jira";    ws["C14"] = "TEC26"
ws["B15"] = "ID Propuesta";     ws["C15"] = "Migración Preselección — 20-may-2026"

for r in range(8, 16):
    ws.cell(row=r, column=2).font = Font(name=FONT, size=10, bold=True, color="000000")
    ws.cell(row=r, column=2).fill = fill(LT_GREY)
    ws.cell(row=r, column=3).font = body_font()
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

# Control de versiones
ws["B18"] = "CONTROL DE VERSIONES"
ws["B18"].font = Font(name=FONT, size=12, bold=True, color=LIME)
ws.merge_cells("B18:G18")

cv_hdr = ["Versión", "Fecha", "Autor", "Cambios", "Estado", "Aprobado por"]
for i, h in enumerate(cv_hdr):
    c = ws.cell(row=19, column=2+i, value=h)
    c.font = hdr_font(); c.fill = fill(BG_DARK); c.alignment = center(); c.border = border

ws.append([])  # row 20 spacer not needed since explicit row use
v1 = ["v1.0", "25-may-2026", "Miroslava Jiménez", "Versión inicial del RAID Log", "Borrador", "PO TecMilenio"]
for i, val in enumerate(v1):
    c = ws.cell(row=20, column=2+i, value=val)
    c.font = body_font(); c.alignment = left_wrap(); c.border = border
    c.fill = fill(LT_GREY if i % 2 == 0 else "FFFFFF")

# Leyenda de semaforización
ws["B23"] = "LEYENDA DE SEMAFORIZACIÓN"
ws["B23"].font = Font(name=FONT, size=12, bold=True, color=LIME); ws.merge_cells("B23:G23")

legend = [
    ("CRÍTICO", RED,    "Impacto severo en cronograma/costo · Mitigación inmediata requerida"),
    ("ALTO",    ORANGE, "Impacto significativo · Plan de mitigación activo"),
    ("MEDIO",   LIME,   "Impacto moderado · Monitoreo permanente"),
    ("BAJO",    BLUE,   "Impacto menor · Aceptable con monitoreo"),
]
for i, (lvl, color, desc) in enumerate(legend):
    r = 24 + i
    c1 = ws.cell(row=r, column=2, value=lvl)
    c1.font = Font(name=FONT, size=10, bold=True, color=BG_DARK)
    c1.fill = fill(color); c1.alignment = center(); c1.border = border
    c2 = ws.cell(row=r, column=3, value=desc)
    c2.font = body_font(); c2.alignment = left_wrap(); c2.border = border
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

# Footer
ws["B30"] = "© 2026 Mobiik   |   coding for the future   |   Confidencial"
ws["B30"].font = Font(name=FONT, size=9, italic=True, color=GREY)
ws.merge_cells("B30:G30")


# ─── HELPER PARA HOJAS DE TABLA ───────────────────────────────────
def write_table_sheet(ws_name, title, headers, rows, col_widths, severity_col=None):
    ws2 = wb.create_sheet(ws_name)
    ws2.sheet_view.showGridLines = False
    # Column widths
    ws2.column_dimensions['A'].width = 2
    for i, w in enumerate(col_widths):
        ws2.column_dimensions[get_column_letter(2 + i)].width = w

    # Title strip (dark)
    n_cols = len(headers)
    last_col = get_column_letter(1 + n_cols)
    for col in range(1, n_cols + 2):
        ws2.cell(row=1, column=col).fill = fill(BG_DARK)
        ws2.cell(row=2, column=col).fill = fill(BG_DARK)
        ws2.cell(row=3, column=col).fill = fill(BG_DARK)
    ws2.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1+n_cols)
    ws2["B2"] = title
    ws2["B2"].font = title_font(); ws2["B2"].alignment = Alignment(horizontal="left", vertical="center")

    # Lime bar
    for col in range(2, 2+n_cols):
        ws2.cell(row=4, column=col).fill = fill(LIME)
    ws2.row_dimensions[4].height = 3

    # Headers
    for i, h in enumerate(headers):
        c = ws2.cell(row=6, column=2+i, value=h)
        c.font = hdr_font(); c.fill = fill(BG_DARK); c.alignment = center(); c.border = border
    ws2.row_dimensions[6].height = 28

    # Rows
    for ri, row in enumerate(rows):
        r = 7 + ri
        for i, val in enumerate(row):
            c = ws2.cell(row=r, column=2+i, value=val)
            c.font = body_font(); c.alignment = left_wrap(); c.border = border
            c.fill = fill(LT_GREY if ri % 2 == 0 else "FFFFFF")
        ws2.row_dimensions[r].height = 45

        # Severity coloring
        if severity_col is not None:
            sev_cell = ws2.cell(row=r, column=2+severity_col)
            sev = (row[severity_col] or "").upper()
            colormap = {"CRÍTICO": RED, "ALTO": ORANGE, "MEDIO": LIME, "BAJO": BLUE,
                        "ALTA": ORANGE, "MEDIA": LIME, "BAJA": BLUE}
            if sev in colormap:
                sev_cell.fill = fill(colormap[sev])
                sev_cell.font = Font(name=FONT, size=10, bold=True, color=BG_DARK)
                sev_cell.alignment = center()

    # Freeze header
    ws2.freeze_panes = "A7"
    return ws2


# ─── HOJA 2: RIESGOS ──────────────────────────────────────────────
riesgos = [
    ["R-01", "Pérdida de datos durante migración de MS SQL Server",
     "Migración de BD desde IaaS a Azure PaaS puede generar pérdida/corrupción de datos por incompatibilidad de esquemas o interrupciones",
     "Técnico", "Media", "Alto", "ALTO",
     "Backup completo previo a migración. Validación de completitud con conteos y checksums. Plan de rollback documentado.",
     "DevOps (Norberto Galicia)", "Abierto", "Sprint 2"],
    ["R-02", "Hallazgos de seguridad adicionales aparecen durante ejecución",
     "TecMilenio identifica hallazgos no contemplados durante el reescaneo, fuera de los 13 incluidos en alcance",
     "Comercial", "Media", "Medio", "MEDIO",
     "Acuerdo explícito en propuesta: hallazgos adicionales requieren change order. Comunicación temprana al PO.",
     "Scrum Master (Miroslava Jiménez)", "Abierto", "Sprint 3"],
    ["R-03", "Retraso en entrega de accesos por TecMilenio",
     "Cliente no entrega accesos a Azure, GitHub o documentación técnica en los primeros 2 días del Sprint 1",
     "Operativo", "Alta", "Alto", "CRÍTICO",
     "Solicitud formal en kickoff. Escalación inmediata al PO si retraso > 24 hrs. Aplicar cláusula de pausa por dependencia (max 3 días).",
     "Scrum Master / PO TecMilenio", "Abierto", "Sprint 1"],
    ["R-04", "Incompatibilidades .NET Framework con Azure App Service",
     "Versión actual de .NET Framework no es compatible con runtime de App Service",
     "Técnico", "Baja", "Medio", "MEDIO",
     "Validación temprana en Sprint 1. Plan de upgrade controlado dentro de misma familia (4.7.x → 4.8.1).",
     "Arquitecto (Andrés Hernández)", "Abierto", "Sprint 1"],
    ["R-05", "Aplicativo tiene bugs preexistentes no detectados por TecMilenio",
     "Supuesto contractual: cliente declara que validó funcionalmente; si aparecen bugs, atribución comercial compleja",
     "Comercial", "Media", "Medio", "MEDIO",
     "Documentar todo bug encontrado y reportarlo formalmente al PO antes de remediar. Estimar como change order.",
     "QA (César Zúñiga)", "Abierto", "Sprint 2"],
    ["R-06", "Equipo de Pruebas del cliente no disponible en UAT",
     "TecMilenio no asigna equipo de pruebas durante los 3 días de UAT en Semana 3",
     "Operativo", "Media", "Alto", "ALTO",
     "Confirmación formal de equipo 1 semana antes de UAT. Calendario invitado con PO copiado.",
     "Scrum Master", "Abierto", "Sprint 3"],
    ["R-07", "Modernizr / Html5Shiv son librerías deprecadas sin updates",
     "2 de los 13 hallazgos están sobre librerías sin mantenimiento — la mitigación puede requerir reemplazo, no upgrade",
     "Técnico", "Alta", "Bajo", "BAJO",
     "Validar opciones: reemplazo por equivalentes mantenidos o aceptación documentada del riesgo por TecMilenio.",
     "Arquitecto (Andrés Hernández)", "Abierto", "Sprint 3"],
    ["R-08", "Pausa por dependencia > 3 días genera penalización contractual",
     "Cláusula contractual: pausas mayores a 3 días aplican penalización del 80% de tarifa comercial por persona",
     "Comercial", "Media", "Alto", "ALTO",
     "Scrum Master notifica preventivamente dependencias próximas. Documentación de cada pausa con timestamp.",
     "Scrum Master", "Abierto", "Continuo"],
]

write_table_sheet(
    ws_name="1. Riesgos",
    title="🔴  RIESGOS DEL PROYECTO",
    headers=["ID", "Riesgo", "Descripción", "Categoría", "Probabilidad", "Impacto", "Severidad", "Mitigación", "Responsable", "Estado", "Sprint"],
    rows=riesgos,
    col_widths=[8, 35, 50, 15, 13, 12, 13, 50, 25, 12, 12],
    severity_col=6
)

# ─── HOJA 3: SUPUESTOS ────────────────────────────────────────────
supuestos = [
    ["A-01", "Cliente validó funcionalmente el aplicativo y no tiene bugs conocidos",
     "Si aparecen bugs durante migración, requieren change order", "TecMilenio", "Confirmado"],
    ["A-02", "Documentación técnica del aplicativo disponible y actualizada",
     "Bloquea inicio de análisis técnico si no se cumple", "TecMilenio", "Pendiente"],
    ["A-03", "Acceso a repositorios GitHub y pipelines existentes",
     "Bloquea configuración de CI/CD si no se cumple", "TecMilenio", "Pendiente"],
    ["A-04", "Se conserva la familia tecnológica actual (sin upgrade a familias superiores)",
     "Upgrades a nuevas familias requieren change order", "Ambos", "Confirmado"],
    ["A-05", "Provisión y configuración de infraestructura productiva la realiza TecMilenio",
     "Mobiik solo provisiona No-Prod y acompaña en Go-Live", "TecMilenio", "Confirmado"],
    ["A-06", "Validación funcional completa post-liberación es responsabilidad del cliente",
     "Mobiik garantiza migración técnica + validación smoke", "TecMilenio", "Confirmado"],
    ["A-07", "Sin pruebas de pentesting ni pruebas de carga (fuera de alcance)",
     "Si se requieren, change order separado", "Ambos", "Confirmado"],
    ["A-08", "Costos de infraestructura Azure los cubre TecMilenio",
     "Mobiik no incluye licenciamiento ni hardware", "TecMilenio", "Confirmado"],
    ["A-09", "Garantía técnica de 30 días naturales post-aceptación",
     "Mobiik corrige defectos atribuibles sin costo durante este periodo", "Mobiik", "Confirmado"],
    ["A-10", "Servicios prestados de manera remota (sin viáticos contemplados)",
     "Viáticos requieren autorización previa de TecMilenio", "Ambos", "Confirmado"],
]

write_table_sheet(
    ws_name="2. Supuestos",
    title="🟡  SUPUESTOS DEL PROYECTO",
    headers=["ID", "Supuesto", "Implicación si no se cumple", "Responsable", "Estado"],
    rows=supuestos,
    col_widths=[8, 60, 45, 18, 15],
)

# ─── HOJA 4: ISSUES ───────────────────────────────────────────────
issues = [
    # vacío al inicio - placeholder para que el equipo lo llene
]
ws_iss = write_table_sheet(
    ws_name="3. Issues",
    title="🟠  ISSUES (Problemas Activos)",
    headers=["ID", "Issue", "Descripción", "Severidad", "Sprint", "Reportado por", "Asignado a", "Estado", "Resolución"],
    rows=[
        ["I-XX", "(Plantilla — sin issues al inicio)",
         "Esta hoja se llenará conforme aparezcan issues durante la ejecución. Mantener formato: I-01, I-02, ...",
         "MEDIO", "—", "—", "—", "Plantilla", "—"]
    ],
    col_widths=[8, 30, 50, 13, 12, 22, 22, 14, 40],
    severity_col=3
)

# ─── HOJA 5: DEPENDENCIAS ─────────────────────────────────────────
deps = [
    ["D-01", "Accesos a repositorios GitHub", "Externa", "TecMilenio (Encargado de Infraestructura)",
     "Sprint 1 (primeros 2 días)", "CRÍTICO", "Bloquea análisis técnico y configuración de CI/CD",
     "Abierto", "Solicitado en kickoff"],
    ["D-02", "Suscripción Azure y permisos para crear recursos", "Externa", "TecMilenio (Encargado de Infraestructura)",
     "Sprint 1 (primeros 2 días)", "CRÍTICO", "Bloquea preparación de entornos completa (App Service, Key Vault, SQL)",
     "Abierto", "Solicitado en kickoff"],
    ["D-03", "Documentación técnica actual del aplicativo Preselección", "Externa", "TecMilenio (Arquitectura)",
     "Sprint 1 (primer día)", "CRÍTICO", "Bloquea análisis técnico y identificación de dependencias",
     "Abierto", "Solicitado en kickoff"],
    ["D-04", "Encargado de Infraestructura disponible para configuración", "Externa", "TecMilenio (Infraestructura)",
     "Sprint 1-2", "ALTO", "Configuración de Azure y permisos requiere su intervención activa",
     "Abierto", "Calendarizar en kickoff"],
    ["D-05", "Confirmación formal de los 13 hallazgos en alcance", "Externa", "TecMilenio (Seguridad)",
     "Sprint 2", "ALTO", "Confirmar que no hay hallazgos adicionales antes de iniciar mitigaciones",
     "Abierto", "Reunión Sprint 2"],
    ["D-06", "Equipo de Pruebas del cliente disponible para UAT (3 días)", "Externa", "TecMilenio (QA)",
     "Sprint 3, Semana 3 (L-X)", "MEDIO", "UAT no puede ejecutarse sin disponibilidad del equipo",
     "Abierto", "Confirmar 1 semana antes"],
    ["D-07", "Provisión y configuración de infraestructura productiva en Azure", "Externa", "TecMilenio (Infraestructura)",
     "Sprint 3 / Pre Go-Live", "MEDIO", "Bloquea Go-Live productivo",
     "Abierto", "Coordinar previo a Go-Live"],
    ["D-08", "Product Owner disponible para validación y toma de decisiones", "Externa", "TecMilenio (Product Owner)",
     "Continuo", "INFORMATIVA", "Decisiones de alcance, priorización y aceptación de entregables",
     "Abierto", "Sesiones agendadas"],
    ["D-09", "Validación funcional completa post-liberación", "Externa", "TecMilenio (Negocio)",
     "Post Go-Live", "INFORMATIVA", "Responsabilidad del cliente — Mobiik solo valida técnicamente",
     "Abierto", "Plan acordado"],
]
write_table_sheet(
    ws_name="4. Dependencias",
    title="🔵  DEPENDENCIAS (Internas y Externas)",
    headers=["ID", "Dependencia", "Tipo", "Responsable", "Cuándo requerida", "Criticidad", "Impacto si no se cumple", "Estado", "Notas"],
    rows=deps,
    col_widths=[8, 35, 12, 28, 22, 14, 40, 12, 25],
    severity_col=5
)

# ─── HOJA 6: DASHBOARD RESUMEN ────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False
ws_dash.column_dimensions['A'].width = 2
for col in 'BCDEFGH':
    ws_dash.column_dimensions[col].width = 16

# Dark title
for col in range(1, 9):
    ws_dash.cell(row=1, column=col).fill = fill(BG_DARK)
    ws_dash.cell(row=2, column=col).fill = fill(BG_DARK)
    ws_dash.cell(row=3, column=col).fill = fill(BG_DARK)
ws_dash.merge_cells("B2:H2")
ws_dash["B2"] = "DASHBOARD RAID"
ws_dash["B2"].font = title_font()
ws_dash["B2"].alignment = Alignment(horizontal="left", vertical="center")

for col in range(2, 9):
    ws_dash.cell(row=4, column=col).fill = fill(LIME)
ws_dash.row_dimensions[4].height = 3

# KPI cards
kpis = [
    ("RIESGOS\nABIERTOS",     '=COUNTIF(\'1. Riesgos\'!K7:K100,"Abierto")', RED),
    ("RIESGOS\nCRÍTICOS",     '=COUNTIF(\'1. Riesgos\'!H7:H100,"CRÍTICO")', RED),
    ("SUPUESTOS\nPENDIENTES", '=COUNTIF(\'2. Supuestos\'!F7:F100,"Pendiente")', ORANGE),
    ("ISSUES\nACTIVOS",       '=COUNTIF(\'3. Issues\'!I7:I100,"Abierto")', ORANGE),
    ("DEPENDENCIAS\nCRÍTICAS",'=COUNTIF(\'4. Dependencias\'!G7:G100,"CRÍTICO")', RED),
    ("BLOQUEADORES\nCLIENTE", '=COUNTIF(\'4. Dependencias\'!D7:D100,"Externa")', ORANGE),
]
for i, (label, formula, color) in enumerate(kpis):
    col = 2 + i
    # Label
    c = ws_dash.cell(row=6, column=col, value=label)
    c.font = Font(name=FONT, size=9, bold=True, color=WHITE)
    c.fill = fill(BG_PANEL); c.alignment = center(); c.border = border
    ws_dash.row_dimensions[6].height = 35
    # Big number
    c2 = ws_dash.cell(row=7, column=col, value=formula)
    c2.font = Font(name=FONT, size=28, bold=True, color=color)
    c2.fill = fill(BG_DARK); c2.alignment = center(); c2.border = border
    ws_dash.row_dimensions[7].height = 55

# Distribución por severidad
ws_dash["B10"] = "DISTRIBUCIÓN DE RIESGOS POR SEVERIDAD"
ws_dash["B10"].font = Font(name=FONT, size=12, bold=True, color=LIME)
ws_dash.merge_cells("B10:H10")

sev_hdr = ["Severidad", "Cantidad"]
for i, h in enumerate(sev_hdr):
    c = ws_dash.cell(row=11, column=2+i, value=h)
    c.font = hdr_font(); c.fill = fill(BG_DARK); c.alignment = center(); c.border = border

sev_rows = [
    ("CRÍTICO", RED),
    ("ALTO",    ORANGE),
    ("MEDIO",   LIME),
    ("BAJO",    BLUE),
]
for i, (lvl, color) in enumerate(sev_rows):
    r = 12 + i
    c1 = ws_dash.cell(row=r, column=2, value=lvl)
    c1.font = Font(name=FONT, size=10, bold=True, color=BG_DARK)
    c1.fill = fill(color); c1.alignment = center(); c1.border = border
    c2 = ws_dash.cell(row=r, column=3, value=f'=COUNTIF(\'1. Riesgos\'!H7:H100,"{lvl}")')
    c2.font = Font(name=FONT, size=12, bold=True, color="000000")
    c2.alignment = center(); c2.border = border; c2.fill = fill(LT_GREY)

# Footer
ws_dash["B30"] = "© 2026 Mobiik   |   coding for the future"
ws_dash["B30"].font = Font(name=FONT, size=9, italic=True, color=GREY)
ws_dash.merge_cells("B30:H30")

# Reorder: Portada, Dashboard, then RAID sheets
wb._sheets = [wb["Portada"], wb["Dashboard"], wb["1. Riesgos"], wb["2. Supuestos"], wb["3. Issues"], wb["4. Dependencias"]]

# Save
out = "./entregables/RAID Log - Migración Preselección.xlsx"
wb.save(out)
print(f"✓ Saved: {out}")
print(f"  Sheets: {wb.sheetnames}")
